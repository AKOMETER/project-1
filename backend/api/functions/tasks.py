from celery import shared_task
from api.models import PhoneNumber
import logging
from django.core.mail import send_mail, EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
import requests, json
import threading
from django.db import transaction
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@shared_task
def send_message_to_facebook_excel_normal(
    excel, template_name, user_id, phone_number_id, bearer_token
):
    try:
        logger.info("Excel Data: %s", excel)
        results = []
        # excel_data = pd.read_json(excel_data_json)
        for value in excel:
            raw_number = str(value).replace(" ", "")
            if raw_number and (raw_number.startswith("91")) and (len(raw_number) == 12):
                raw_number = "+" + raw_number
            elif raw_number and raw_number.startswith("0"):
                raw_number = "+91" + raw_number[1:]
            print(raw_number)

            if raw_number:
                # PhoneNumber.objects.get_or_create(number=raw_number, user_id=user_id)

                data = {
                    "messaging_product": "whatsapp",
                    "recipient_type": "individual",
                    "to": raw_number,
                    "type": "template",
                    "template": {"name": template_name, "language": {"code": "en"}},
                }

                url = f"https://graph.facebook.com/v18.0/{phone_number_id}/messages"
                headers = {
                    "Authorization": "Bearer " + bearer_token,
                    "Content-Type": "application/json",
                }

                try:
                    response = requests.post(url, headers=headers, json=data)
                    response_data = response.json()
                    results.append(response_data)
                except json.JSONDecodeError:
                    pass
        return "None"

        # return results
    except Exception as e:
        # logger.exception("An error occurred in send_message_to_facebook task: %s", e)
        # print("dscjbjhb")

        raise e


@shared_task
def send_message_to_facebook_excel_images(
    excel,
    template_name,
    user_id,
    phone_number_id,
    bearer_token,
    image_link_get,
    template_format,
):
    def process_number(
        value,
        user_id,
        template_name,
        template_format,
        phone_number_id,
        bearer_token,
        results,
    ):
        raw_number = str(value).replace(" ", "")
        if raw_number and (raw_number.startswith("91")) and (len(raw_number) == 12):
            raw_number = "+" + raw_number
        elif raw_number and raw_number.startswith("0"):
            raw_number = "+91" + raw_number[1:]
        # print(raw_number)

        if raw_number:
            # with transaction.atomic():
            # PhoneNumber.objects.get_or_create(number=raw_number, user_id=user_id)

            data = {
                "messaging_product": "whatsapp",
                "recipient_type": "individual",
                "to": raw_number,
                "type": "template",
                "template": {
                    "name": template_name,
                    "language": {"code": "en"},
                    **(
                        {
                            "components": [
                                {
                                    "type": "header",
                                    "parameters": [
                                        {
                                            "type": template_format,
                                            template_format: {
                                                "link": image_link_get,
                                                **(
                                                    {"filename": f"{template_name}.pdf"}
                                                    if template_format == "document"
                                                    else {}
                                                ),
                                            },
                                        }
                                    ],
                                }
                            ],
                        }
                        if template_format != "text"
                        else {}
                    ),
                },
            }
            url = f"https://graph.facebook.com/v18.0/{phone_number_id}/messages"
            headers = {
                "Authorization": "Bearer " + bearer_token,
                "Content-Type": "application/json",
            }

            try:
                response = requests.post(url, headers=headers, json=data)
                response_data = response.json()
                results.append(response_data)
                # print(results)
            except json.JSONDecodeError:
                print("JSON Decode Error")

    try:
        # logger.info("Excel Data: %s", excel)
        results = []
        threads = []

        for value in excel:
            thread = threading.Thread(
                target=process_number,
                args=(
                    value,
                    user_id,
                    template_name,
                    template_format,
                    phone_number_id,
                    bearer_token,
                    results,
                ),
            )
            threads.append(thread)
            thread.start()

        for thread in threads:
            thread.join()

        return results
    except Exception as e:
        logger.exception("An error occurred in send_message_to_facebook task: %s", e)
        raise e


@shared_task
def send_message_to_facebook_array(
    numbers,
    template_name,
    user_id,
    phone_number_id,
    bearer_token,
    image_link_get,
    template_format,
):
    def process_number(
        value,
        user_id,
        template_name,
        template_format,
        phone_number_id,
        bearer_token,
        results,
    ):
        raw_number = str(value).replace(" ", "")
        if raw_number and (raw_number.startswith("91")) and (len(raw_number) == 12):
            raw_number = "+" + raw_number
        elif raw_number and raw_number.startswith("0"):
            raw_number = "+91" + raw_number[1:]
        # print(raw_number)

        if raw_number:
            # with transaction.atomic():
            # PhoneNumber.objects.get_or_create(number=raw_number, user_id=user_id)

            data = {
                "messaging_product": "whatsapp",
                "recipient_type": "individual",
                "to": raw_number,
                "type": "template",
                "template": {
                    "name": template_name,
                    "language": {"code": "en"},
                    **(
                        {
                            "components": [
                                {
                                    "type": "header",
                                    "parameters": [
                                        {
                                            "type": template_format,
                                            template_format: {
                                                "link": image_link_get,
                                                **(
                                                    {"filename": f"{template_name}.pdf"}
                                                    if template_format == "document"
                                                    else {}
                                                ),
                                            },
                                        }
                                    ],
                                }
                            ],
                        }
                        if template_format != "text"
                        else {}
                    ),
                },
            }
            url = f"https://graph.facebook.com/v18.0/{phone_number_id}/messages"
            headers = {
                "Authorization": "Bearer " + bearer_token,
                "Content-Type": "application/json",
            }

            try:
                response = requests.post(url, headers=headers, json=data)
                response_data = response.json()
                results.append(response_data)
                # print(results)
            except json.JSONDecodeError:
                print("JSON Decode Error")

    try:
        # logger.info("Excel Data: %s", excel)
        results = []
        threads = []

        for value in numbers:
            thread = threading.Thread(
                target=process_number,
                args=(
                    value,
                    user_id,
                    template_name,
                    template_format,
                    phone_number_id,
                    bearer_token,
                    results,
                ),
            )
            threads.append(thread)
            thread.start()

        for thread in threads:
            thread.join()

        return results
    except Exception as e:
        logger.exception("An error occurred in send_message_to_facebook task: %s", e)
        raise e


@shared_task
def send_personalized_messages(
    excel_file_path, template_name, phone_number_id, bearer_token, user_id
):
    workbook = load_workbook(excel_file_path, read_only=True)
    worksheet = workbook.active
    results = []

    for row in worksheet.iter_rows(values_only=True):
        raw_number = str(row[0])
        name = str(row[1])

        if raw_number.startswith("0"):
            raw_number = "+91" + raw_number[1:]

        if not raw_number.startswith("+91"):
            raw_number = "+91" + raw_number

        # PhoneNumber.objects.get_or_create(number=raw_number, user_id=user_id)

        data = {
            "messaging_product": "whatsapp",
            "recipient_type": "individual",
            "to": raw_number,
            "type": "template",
            "template": {
                "name": template_name,
                "components": [
                    {"type": "HEADER", "parameters": [{"type": "text", "text": name}]}
                ],
                "language": {"code": "en"},
            },
        }

        url = f"https://graph.facebook.com/v18.0/{phone_number_id}/messages"
        headers = {
            "Authorization": f"Bearer {bearer_token}",
            "Content-Type": "application/json",
        }

        try:
            response = requests.post(url, headers=headers, json=data)
            response_data = response.json()
            results.append(response_data)

        except json.JSONDecodeError:
            results.append({"error": "Invalid JSON data"})
    return results


@shared_task
def send_message_to_facebook_custom(
    message,
    numbers,
    phone_number_id,
    bearer_token,
):
    def process_number(message, value, phone_number_id, bearer_token, results):
        raw_number = str(value).replace(" ", "")
        if raw_number and (raw_number.startswith("91")) and (len(raw_number) == 12):
            raw_number = "+" + raw_number
        elif raw_number and raw_number.startswith("0"):
            raw_number = "+91" + raw_number[1:]
        # print(raw_number)

        if raw_number:
            data = {
                "messaging_product": "whatsapp",
                "recipient_type": "individual",
                "to": raw_number,
                "type": "text",
                "text": {"preview_url": True, "body": message},
            }
            url = f"https://graph.facebook.com/v18.0/{phone_number_id}/messages"
            headers = {
                "Authorization": "Bearer " + bearer_token,
                "Content-Type": "application/json",
            }

            try:
                response = requests.post(url, headers=headers, json=data)
                response_data = response.json()
                results.append(response_data)
                print(response_data)
            except json.JSONDecodeError:
                print("JSON Decode Error")

    try:
        # logger.info("Excel Data: %s", excel)
        results = []
        threads = []

        for value in numbers:
            thread = threading.Thread(
                target=process_number,
                args=(message, value, phone_number_id, bearer_token, results),
            )
            threads.append(thread)
            thread.start()

        for thread in threads:
            thread.join()

        return results
    except Exception as e:
        logger.exception("An error occurred in send_message_to_facebook task: %s", e)
        raise e


@shared_task
def send_email(email, password):
    my_subject = "Whatsapp Module Generated Password"
    context = {
        "password": password,
    }
    recipient_list = [email]
    html_message = render_to_string("password.html", context)
    plain_message = strip_tags(html_message)
    message = EmailMultiAlternatives(
        subject=my_subject,
        body=plain_message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=recipient_list,
    )
    message.attach_alternative(html_message, "text/html")
    message.send()
    # return None
